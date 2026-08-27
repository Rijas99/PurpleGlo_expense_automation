from __future__ import annotations

import shutil
from pathlib import Path

from openpyxl import load_workbook

from app.constants import EXCEL_DATA_END_ROW, EXCEL_DATA_START_ROW, TEMPLATE_RELATIVE_PATH

_COL = {
    "ref": "C",
    "date": "D",
    "description": "E",
    "category": "F",
    "project_code": "G",
    "amount": "H",
}


def project_code_for_form(rec: dict) -> str:
    code = str(rec.get("project_code") or "").strip()
    if code:
        return code
    return str(rec.get("project_name") or "").strip()



def fill_expense_form(
    receipts: list[dict],
    month: str,
    dest: Path,
    template_path: Path | str | None = None,
    employee_name: str = "Rijas Ali",
) -> Path:
    """Copy the company template and write receipt rows into Sheet1. Layout and logo stay intact."""
    dest = Path(dest)
    dest.parent.mkdir(parents=True, exist_ok=True)
    src = Path(template_path) if template_path else Path(TEMPLATE_RELATIVE_PATH)
    shutil.copy2(src, dest)

    wb = load_workbook(dest)
    ws = wb["Sheet1"]
    ws["D9"] = employee_name
    ws["D10"] = month

    capacity = EXCEL_DATA_END_ROW - EXCEL_DATA_START_ROW + 1
    rows = list(receipts)
    if len(rows) > capacity:
        extra = len(rows) - capacity
        ws.insert_rows(EXCEL_DATA_END_ROW + 1, extra)
        last_data = EXCEL_DATA_END_ROW + extra
        # Copy borders/number format from the last original data row onto new rows
        template_row = EXCEL_DATA_END_ROW
        for offset in range(1, extra + 1):
            new_row = EXCEL_DATA_END_ROW + offset
            for col in ("C", "D", "E", "F", "G", "H"):
                src_cell = ws[f"{col}{template_row}"]
                dst_cell = ws[f"{col}{new_row}"]
                dst_cell.number_format = src_cell.number_format
                dst_cell.border = src_cell.border.copy()
                dst_cell.alignment = src_cell.alignment.copy()
                dst_cell.font = src_cell.font.copy()
        ws[f"H{last_data + 1}"] = f"=SUM(H{EXCEL_DATA_START_ROW}:H{last_data})"
    else:
        last_data = EXCEL_DATA_END_ROW
        # Keep original TOTAL formula on H34
        if ws["H34"].value is None:
            ws["H34"] = f"=SUM(H{EXCEL_DATA_START_ROW}:H{EXCEL_DATA_END_ROW})"

    # Clear existing data rows in the write range
    write_end = EXCEL_DATA_START_ROW + max(len(rows), 0) - 1
    clear_end = max(write_end, EXCEL_DATA_END_ROW)
    if len(rows) > capacity:
        clear_end = last_data
    for r in range(EXCEL_DATA_START_ROW, clear_end + 1):
        for col in _COL.values():
            ws[f"{col}{r}"] = None

    for i, rec in enumerate(rows):
        r = EXCEL_DATA_START_ROW + i
        ws[f"C{r}"] = int(rec.get("ref") or i + 1)
        ws[f"D{r}"] = rec.get("date") or ""
        ws[f"E{r}"] = rec.get("description") or ""
        ws[f"F{r}"] = rec.get("category") or ""
        ws[f"G{r}"] = project_code_for_form(rec)
        try:
            ws[f"H{r}"] = float(rec.get("amount") or 0)
        except (TypeError, ValueError):
            ws[f"H{r}"] = rec.get("amount") or 0

    wb.save(dest)
    wb.close()
    return dest
