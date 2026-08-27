from pathlib import Path

import pytest
from openpyxl import load_workbook

from app.constants import EXCEL_DATA_START_ROW, TEMPLATE_RELATIVE_PATH
from app.excel_fill import fill_expense_form


@pytest.fixture
def template_path():
    path = Path(TEMPLATE_RELATIVE_PATH)
    if not path.exists():
        pytest.skip("Expense Form template not found")
    return path


def test_fill_writes_month_name_and_rows(tmp_path, template_path):
    dest = tmp_path / "out.xlsx"
    receipts = [
        {
            "ref": 1,
            "date": "15-Aug",
            "description": "Lunch",
            "category": "Food & Beverages",
            "project_code": "250909-PDS-303",
            "amount": 40.0,
        },
        {
            "ref": 2,
            "date": "16-Aug",
            "description": "Parking",
            "category": "Parking",
            "project_code": "250909-PDS-303",
            "amount": 12.5,
        },
    ]
    fill_expense_form(
        receipts,
        month="Aug 2026",
        dest=dest,
        template_path=template_path,
        employee_name="Rijas Ali",
    )
    wb = load_workbook(dest)
    ws = wb["Sheet1"]
    assert ws["D9"].value == "Rijas Ali"
    assert ws["D10"].value == "Aug 2026"
    row = EXCEL_DATA_START_ROW
    assert ws[f"C{row}"].value == 1
    assert ws[f"D{row}"].value == "15-Aug"
    assert ws[f"E{row}"].value == "Lunch"
    assert ws[f"F{row}"].value == "Food & Beverages"
    assert ws[f"G{row}"].value == "250909-PDS-303"
    assert ws[f"H{row}"].value == 40.0
    assert ws[f"C{row + 1}"].value == 2
    assert ws[f"H{row + 1}"].value == 12.5
    assert ws["H34"].value == "=SUM(H13:H33)"
    assert ws["C38"].value == "Line Manager Approval"
    assert len(ws._images) >= 1


def test_fill_does_not_write_project_name_column(tmp_path, template_path):
    dest = tmp_path / "out.xlsx"
    fill_expense_form(
        [
            {
                "ref": 1,
                "date": "01-Jan",
                "description": "Hotel",
                "category": "Hotel Booking",
                "project_code": "X",
                "project_name": "Should Not Appear",
                "amount": 100,
            }
        ],
        month="Jan 2026",
        dest=dest,
        template_path=template_path,
    )
    wb = load_workbook(dest)
    ws = wb["Sheet1"]
    # Header row stays REF DATE DESCRIPTION CATEGORY PROJECT CODE AMOUNT
    assert ws["C12"].value == "REF"
    assert ws["H12"].value == "AMOUNT"
    assert "Should Not Appear" not in [
        ws.cell(row=13, column=c).value for c in range(1, 12)
    ]


def test_fill_uses_project_name_when_code_missing(tmp_path, template_path):
    dest = tmp_path / "out.xlsx"
    fill_expense_form(
        [
            {
                "ref": 1,
                "date": "01-Jan",
                "description": "Hotel",
                "category": "Hotel Booking",
                "project_code": "",
                "project_name": "Site Alpha",
                "amount": 100,
            }
        ],
        month="Jan 2026",
        dest=dest,
        template_path=template_path,
    )
    wb = load_workbook(dest)
    ws = wb["Sheet1"]
    assert ws["G13"].value == "Site Alpha"
